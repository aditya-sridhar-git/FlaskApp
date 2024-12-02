from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook
import csv
import math
from datetime import datetime
import os
#import mysql.connector as sqltor

app = Flask(__name__)

# Path to store the Excel file
EXCEL_FILE = 'travel.xlsx'

# Create a new Excel workbook if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['Starting Destinaiton','Ending Destination'])
    wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('indexTest.html',title='EL Project')

@app.route('/submit_data', methods=['POST'])
def submit_data():
    try:
         
        data = request.get_json()  # Get the data from the form
        #print(data)
        
        '''email_id = data.get('email_id')'''
        startDestination= data.get('startDestination')
        endDestination = data.get('endDestination')
        
        '''distance_travelled = data.get('distance_travelled')
        total_travel_time = data.get('total_travel_time')'''
        

        if not startDestination or not endDestination :
            return jsonify({'success': False, 'message': 'apptest: Both fields are required!'})

        # Open the Excel file to append data
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        row_count = ws.max_row
        ws.append([startDestination,endDestination])
        wb.save(EXCEL_FILE)

        return jsonify({'success': True, 'message': 'Data successfully added to Excel file!'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
        
# File to store the data
#CSV_FILE = 

@app.route('/save-opted-path', methods=['POST'])
def save_opted_path():
    data = request.get_json()
    total_distance = data.get("totalDistance")

    if total_distance:
        # Append the opted path distance to the CSV
        with open(CSV_FILE, mode='a', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Opted Path", f"{total_distance} km"])
        return jsonify({"message": "Opted path saved successfully!"}), 200
    else:
        return jsonify({"error": "Invalid data provided"}), 400
        
# Function to calculate distance using the Haversine formula
def haversine(lat1, lon1, lat2, lon2):
    R = 6371  # Radius of the Earth in kilometers
    to_rad = lambda x: x * math.pi / 180
    dlat = to_rad(lat2 - lat1)
    dlon = to_rad(lon2 - lon1)
    a = math.sin(dlat / 2) ** 2 + math.cos(to_rad(lat1)) * math.cos(to_rad(lat2)) * math.sin(dlon / 2) ** 2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c  # Distance in kilometers
    
# Route to save location data to CSV and calculate distance
@app.route('/save_location_data', methods=['POST'])
def save_location_data():
    data = request.json
    starting_location = data.get('startingLocation')
    tracking_data = data.get('trackingData')

    filename = "location_data.csv"
    total_distance = 0.0

    with open(filename, mode='a', newline='') as file:
        writer = csv.writer(file)
        
        # Write headers if file is empty
        if file.tell() == 0:
            writer.writerow(["Latitude", "Longitude", "Timestamp", "Type"])

        # Write starting location
        if starting_location:
            writer.writerow([
                starting_location['latitude'],
                starting_location['longitude'],
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "Start"
            ])

        # Calculate distances between points
        prev_location = starting_location
        for data_point in tracking_data:
            current_location = {
                'latitude': data_point['latitude'],
                'longitude': data_point['longitude']
            }
            # Calculate distance using haversine
            distance = haversine(
                float(prev_location['latitude']),
                float(prev_location['longitude']),
                float(current_location['latitude']),
                float(current_location['longitude'])
            )
            total_distance += distance
            prev_location = current_location

            # Write intermediate location
            writer.writerow([
                data_point['latitude'],
                data_point['longitude'],
                data_point['timestamp'],
                "Intermediate"
            ])

        # Write ending location
        if tracking_data:
            ending_location = tracking_data[-1]
            writer.writerow([
                ending_location['latitude'],
                ending_location['longitude'],
                ending_location['timestamp'],
                "End"
            ])

    # Handle the case where no movement occurred
    if total_distance == 0:
        total_distance = 0  # Ensure the distance is explicitly 0 in the response

    return jsonify({"status": "success", "message": "Location data saved successfully!", "total_distance": total_distance})

'''
def writeToDatabase(EXCEL_FILE):
    
    
    
    cnx = sqltor.connect(user = "root",host = "localhost",password = "sql123")
    cursor = cnx.cursor()
    query = "use el"
    cursor.execute(query)
    
    query = "select count(*) from data"
    cursor.execute(query)
    data = cursor.fetchall()
    row_count = data[0][0]
    
    f = open(EXCEL_FILE+".csv","a")
    readerOb= csv.reader(f)
    
    for i in readerOb():
        query = "insert into data values(%s,%s,%s)"
        cursor.execute(query,(i[0],i[1],i[2]))
''' 
        


if __name__ == '__main__':
    app.run(debug=True)