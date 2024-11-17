from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)

# Path to store the Excel file
EXCEL_FILE = 'travel_destinations.xlsx'

# Create a new Excel workbook if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['Starting Destination', 'Ending Destination'])
    wb.save(EXCEL_FILE)

@app.route('/')
def index():
    return render_template('index.html',title='EL Project')

@app.route('/submit_data', methods=['POST'])
def submit_data():
    try:
         
        
        data = request.get_json()  # Get the data from the form
        print(data)
        starting_destination = data.get('startingDestination')
        ending_destination = data.get('endingDestination')

        if not starting_destination or not ending_destination:
            return jsonify({'success': False, 'message': 'Both fields are required!'})

        # Open the Excel file to append data
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([starting_destination, ending_destination])
        wb.save(EXCEL_FILE)

        return jsonify({'success': True, 'message': 'Data successfully added to Excel file!'})

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

if __name__ == '__main__':
    app.run(debug=True)